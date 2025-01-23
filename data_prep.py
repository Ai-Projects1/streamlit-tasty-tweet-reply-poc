import os
import json
from openpyxl import load_workbook
from io import BytesIO
from PIL import Image as PILImage

# Input file and output folder
input_filepath = 'assets/Tasty Twitter Reply Data.xlsx'  # Path to the Excel file
sheet_name = 'DATA'  # Sheet name containing images
output_folder = 'assets/extracted_images'  # Folder to save the extracted images
jsonl_filename = 'assets/extracted_images_data.jsonl'  # Path for the jsonl file

# Create the output folder if it doesn't exist
os.makedirs(output_folder, exist_ok=True)

# Load the workbook and sheet
workbook = load_workbook(input_filepath)
sheet = workbook[sheet_name]

# Initialize data structures
training_data = []
column_4_values = []  # Values from column 4 with row number
column_6_values = []  # Values from column 6 with row number
image_positions = []  # List to store image_row and image_col
images_without_captions = []  # List to store images without captions
main_caption_reply_data = []  # List to store main caption and reply pairs

# Iterate over all images in the sheet
for i, drawing in enumerate(sheet._images, start=1):  # `_images` contains all embedded images
    # Extract the image extension (if available)
    image_extension = drawing.path.split('.')[-1] if drawing.path else 'png'  # Default to 'png' if no extension

    # Get the corresponding caption
    anchor_cell = drawing.anchor
    anchor_position = anchor_cell._from  # Get the anchor's position (this is a Point object)
    image_row = anchor_position.row  # Row number
    image_col = anchor_position.col  # Column number

    # Append image position to the list
    image_positions.append({"image_row": image_row, "image_col": image_col})

    # Initialize caption and subfolder
    caption = None
    subfolder = 'Uncategorized'

    # Process images in column 2 (Main Post)
    if image_col == 2:  # Column C is the 3rd column (index 2)
        caption = sheet.cell(row=image_row + 1, column=4).value  # Caption in column D
        if not caption:
            caption = sheet.cell(row=image_row + 2, column=4).value  # Fallback to next row
        subfolder = 'Main Post'

        # Store the caption for use in column 4 images later
        main_caption = caption

    # Process images in column 4 (Reply)
    elif image_col == 4:  # Column E is the 5th column (index 4)
        caption = sheet.cell(row=image_row + 1, column=6).value  # Caption in column F
        if not caption:
            caption = sheet.cell(row=image_row + 2, column=6).value  # Fallback to next row
        subfolder = 'Reply'

        # If there is a caption in column 2, include it in the `text` key
        main_caption_text = main_caption if 'main_caption' in locals() else "No main caption"

    # Skip processing if no caption is available
    if not caption and image_col != 4:
        images_without_captions.append({"image_name": f"image_{i}.{image_extension}", "image_row": image_row, "image_col": image_col})
        continue

    # Create subfolder if it doesn't exist
    subfolder_path = os.path.join(output_folder, subfolder)
    os.makedirs(subfolder_path, exist_ok=True)

    # Create a file name for the image
    image_name = f"image_{i}.{image_extension}"
    image_path = os.path.join(subfolder_path, image_name)  # Full path for the image file

    # Get the image as a byte stream
    image_data = drawing.ref  # BytesIO object

    # Open the image with Pillow and save it as a standard format (e.g., PNG)
    with PILImage.open(image_data) as img:
        img.save(image_path)  # Save the image in the correct format

    print(f"Saved: {image_path}")

    # For column 4 images, add to training_data
    if caption and (image_col == 4) and (caption != main_caption_text):
        input_text = f"Generate {subfolder} describing image." if not main_caption_text else f"Generate {subfolder} that is connected to the image and responding to the main caption. Main caption is {main_caption_text}"
        input_text = input_text.replace('\n', '')
        caption = caption.replace('\n', '')

        json_obj = {
            "contents": [
                {
                    "role": "user",
                    "parts": [
                        {
                            "fileData": {
                                "mimeType": f"image/{image_extension}",
                                "fileUri": f"gs://b-bucket/images/{subfolder}/{image_name}"
                            }
                        },
                        {
                            "text": ''
                        }
                    ]
                },
                {
                    "role": "model",
                    "parts": [
                        {
                            "text": caption
                        }
                    ]
                }
            ]
        }
        training_data.append(json_obj)

        # Collect main caption and reply for new JSON file
        if 'main_caption' in locals() and 'caption' in locals() and main_caption and caption:
            main_caption_reply_data.append({
                "Main caption": main_caption_text,
                "Reply": caption
            })

    # Track images without captions
    if not caption:
        images_without_captions.append({"image_name": image_name, "image_row": image_row, "image_col": image_col})


# Extract all values from columns 4 and 6 with their row numbers
for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
    if row[3]:  # Check if column 4 (index 3) is not empty
        column_4_values.append({"caption": row[3], "row_number": row_number})
    if row[5]:  # Check if column 6 (index 5) is not empty
        column_6_values.append({"caption": row[5], "row_number": row_number})

# Save all data to JSON files
with open('assets/training_data.json', 'w', encoding='utf-8') as json_file:
    json.dump(training_data, json_file, ensure_ascii=False)

with open('assets/column_4_values.json', 'w') as json_file:
    json.dump(column_4_values, json_file)

with open('assets/column_6_values.json', 'w') as json_file:
    json.dump(column_6_values, json_file)

with open('assets/image_positions.json', 'w') as json_file:
    json.dump(image_positions, json_file)

with open('assets/images_without_captions.json', 'w') as json_file:
    json.dump(images_without_captions, json_file)

# Save main caption and reply data to new JSON file
with open('assets/main_caption_reply.json', 'w', encoding='utf-8') as json_file:
    json.dump(main_caption_reply_data, json_file, ensure_ascii=False)

print("Image extraction, caption saving, and JSON creation complete!")
